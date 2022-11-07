# Прочитать сохранённый csv-файл из задания №18 и сохранить данные
# в excel-файл, кроме возраста – столбец с этими данными не нужен.
import csv
import openpyxl

with open("home_13_02.csv", "r", encoding="utf-8") as f1:
    file_reader = list(csv.reader(f1))
    print(file_reader)

wb = openpyxl.Workbook()

wb.create_sheet(title="List_1", index=0)

wb.remove(wb['Sheet'])
print(wb.sheetnames)

sheet = wb['List_1']
print(sheet)

for col_index, coll in enumerate(file_reader):
    for row_index, value in enumerate(coll):
        cell = sheet.cell(row=row_index+1, column=col_index+1)
        cell.value = value

sheet.insert_rows(1, 1)
insert_strig = ["", "person1", "person2", "person3", "person4", "person5"]
for ind, item in enumerate(insert_strig):
    cell = sheet.cell(row=1, column=ind+1)
    cell.value = item
sheet.delete_rows(4, 1)

wb.save("home_13_03.xlsx")