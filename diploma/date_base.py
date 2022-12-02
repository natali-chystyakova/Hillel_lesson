# Написать программу для работы с данными о людях.
# Программа должна уметь загружать данные из файла, сохранять в файл, вводить
# новые записи и производить поиск по существующим записям.
# Программа сохраняет данные о человеке, а именно: ФИО, дата рождения, дата
# смерти (может отсутствовать) и пол. При этом ФИО вводится 3 полями:
# Имя (обязательно), Фамилия и Отчество могут не вводится.
# Программа должна уметь вычислять возраст человека (количество полных лет) на
# основании даты рождения и даты смерти или сегодняшней даты, если дата смерти
# отсутствует. Дата рождения и дата смерти может вводится в формате:
# 12.10.1980
# 11 10 2000
# 01/02/1995
# 3-9-2007
# Поиск может производится по имени, фамилии и отчеству и выдаёт все варианты,
# которые подходят под строку поиска (это может быть имя, или фамилия, или имя
# и фамилия, или только часть имени и т.д.). К примеру, есть такие записи:
# Евгений Крут Михайлович, 12.10.1980, 11.10.2001, m
# Евгения, 12.10.1980, 12.10.2001, f
# Дмитрий Евгеньевич, 10.03.2000, m
#
# При поиске "евген", выдаются такие данные:
# Евгений Крут Михайлович  20 лет, мужчина. Родился: 12.10.1980. Умер: 11.10.2001.
# Евгения 21 год, женщина. Родилась: 12.10.1980. Умерла: 12.10.2001.
# Дмитрий Евгеньевич 22 года, мужчина. Родился: 10.03.2000.
#
# Программа при старте начинает работать с пустой базой данных. Оператор может
# заполнять её, а может при желании загрузить ранее сохранённые данные из файла
# (желательно Excel).
# Когда есть какие-то записи оператор может сохранить их в файл введя его название.
#
# Желательная структура программы:
# в основной части программы находится вечный цикл с меню, что может выбрать оператор;
# сами данные организованы в виде класса в другом файле, который импортируется в файл
# основной части программы, где создаётся объект соответствующего класса перед заходом
# в вечный цикл;
# все пункты меню основной части программы вызывают те или иные методы у созданного
# объекта данных;
# при желании можно в третьем файле создать отдельный класс Person который будет
# импортироваться в файл с данными. Именно в этом классе будет происходить валидация
# введённых данных.
#
# *Все перечисленные описания являются пожеланиями по реализации дипломного проекта и в
# силу тех или иных причин могут быть изменены по желанию студента. Основные требования:
# программа позволяет ввести новые данные о людях;
# производить поиск по уже введённым данным;
# правильно рассчитывать количество полных лет человека на основе даты рождения и даты
# смерти или текущей даты.
import pandas as pd
import glob
import csv
import json
import openpyxl
from datetime import date


class DB():
    #TOTAL_OBJECTS = 0

    def input_data(self):

        while True:
            name = input("Введите имя ")
            if name == '' or not name.isalpha():
                print('Неверный ввод')
                continue
            else:
                break

        while True:
            print("Чтобы пропустить - нажмите 'Enter'")
            surname = input("Введите фамилию ")
            if surname != '' and not surname.isalpha():
                print('Неверный ввод')
                continue
            else:
                break

        while True:
            print("Чтобы пропустить - нажмите 'Enter'")
            otchestvo = input("Введите отчество ")
            if otchestvo != '' and not otchestvo.isalpha():
                print('Неверный ввод')
                continue

            else:
                break

        while True:
            date_of_birth = input(
                "Введите дату рождения в формате:'dd.mm.yyyy', 'dd/mm/yyyy', 'dd mm yyyy', 'd-m-yyyy'")
            if date_of_birth == "":
                print('Неверный ввод')
                continue
            for item in date_of_birth:
                if item not in ("-. 1234567890"):
                    print('Неверный ввод')
                    break
            else:
                break

        while True:
            print("Чтобы пропустить - нажмите 'Enter'")
            date_of_death = input("Введите дату смерти в формате:'dd.mm.yyyy', 'dd/mm/yyyy', 'dd mm yyyy', 'd-m-yyyy'")

            for item in date_of_death:
                if item not in ("-. 1234567890"):
                    print('Неверный ввод')
                    break
            else:
                break



        while True:
            gender = input("введите пол: мужчина - 'm', женщина - 'f' ")
            if len(gender) > 1 or gender not in 'mf':
                print('Неверный ввод')
                continue
            else:
                break

        def fool_age(date_of_birth, date_of_death=""):

            def split_data(data):
                if " " in data:
                    return (data.split(" "))
                elif "." in data:
                    return (data.split("."))
                elif "/" in data:
                    return (data.split("/"))
                elif "-" in data:
                    return (data.split("-"))

            if date_of_death:
                day_b, month_b, year_b = (split_data(date_of_birth))
                day_d, month_d, year_d = (split_data(date_of_death))
                birthday = date(int(year_b), int(month_b), int(day_b))

                death = date(year_d, month_d, day_d)

                age = int(death.strftime("%Y%m%d")) // 10000 - int(birthday.strftime("%Y%m%d")) // 10000
                print(f"Полных лет {age}")
                return age
            else:
                day_b, month_b, year_b = (split_data(date_of_birth))
                birthday = date(int(year_b), int(month_b), int(day_b))

                today = date.today()

                age = int(today.strftime("%Y%m%d")) // 10000 - int(birthday.strftime("%Y%m%d")) // 10000
                print(f"Полных лет {age}")
                return age

        age = fool_age(date_of_birth, date_of_death)

        return DB(name, date_of_birth, gender, age, surname, otchestvo, date_of_death)

    def __init__(self, name, date_of_birth, gender, age, surname="", otchestvo="", date_of_death=""):

        self.name = name
        self.date_of_birth = date_of_birth
        self.gender = gender
        self.age = age
        self.surname = surname
        self.otchestvo = otchestvo
        self.date_of_death = date_of_death
        #DB.TOTAL_OBJECTS += 1


        data_dict = {f'{self.name} {self.surname} {self.otchestvo}': [self.age, self.gender, self.date_of_birth,
                                                                    self.date_of_death]}


        print(data_dict)

        with open("spisok.json", "a", encoding="utf-8") as f:  # открыли файл на запись в формате json
            json.dump(data_dict, f, indent=4)

        line = [self.name, self.surname, self.otchestvo, self.age, self.gender, self.date_of_birth, self.date_of_death]

        #if DB.TOTAL_OBJECTS <= 1:
        with open("spisok.csv", "a", encoding="utf-8", newline='') as f1:
            file_writer = csv.writer(f1)
            file_writer.writerow(line)

    def __str__(self):
        return f'{self.name}{self.otchestvo}{self.surname}{self.gender}{self.age}{self.date_of_birth}{self.date_of_death}'

    def find(self):

        while True:
            name1 = input("Введите поисковый запрос ")
            try:

                with open("spisok.csv", encoding='utf-8') as f:
                    lines = f.readlines()
                    for line in lines:
                        if line.find(name1) != -1:
                            print(name1, ':запись есть')
                            print('строка №', lines.index(line))
                            print('Информация подробно: ', line)


                    break
            except FileNotFoundError:
                print("База данных еще не создана")
                break


    def get_from_file(self):

        # while True:
        #     files = glob.glob("*.csv")
        #     print('файлы для обьединения:', files)
        #     combined = pd.DataFrame()
        #     for file in files:
        #         data = pd.read_csv(file)
        #         data['filename'] = file
        #         combined = pd.concat([combined, data])
        #         combined.to_csv('combined.csv', index=False, sep=';')
        #         print('файлы обьеденены в одну базу данных')
        #     break

        wb2 = openpyxl.load_workbook("for diploma.xlsx")
        print(wb2.sheetnames)

        sheet = wb2.active
        print(sheet)

        rows = sheet.max_row
        cols = sheet.max_column
        print(rows)
        print(cols)
        data = []

        for i in range(2, rows + 1):
            row_data = []
            for j in range(1, cols + 1):
                cell = sheet.cell(row=i, column=j)
                row_data.append(cell.value)
            data.append(row_data)

        print(data)

        with open("spisok.csv", "a", encoding="utf-8", newline='') as f1:
            file_writer = csv.writer(f1)
            for elem in data:
                file_writer.writerow(elem)

    def get_intu_file(self):

        while True:
            with open("spisok.csv", encoding="utf-8") as f:
                file_reader = csv.reader(f)
                data = list(file_reader)
                data11 = open("combined.txt", "w", encoding="utf-8")
                data11.write(f'{data}')
                data11.close()
                print('БД сохранена в формате txt')
            break

    def export_in_json(self):

        while True:
            d = {}
            with open("combined.txt") as file:
                for line in file:
                    key, *value = line.split()
                    d[key] = value
                    print(value)


            with open("combined.json", "w") as f:
                json.dump(value, f, indent=4, ensure_ascii=False )   #
                print('БД готова для передачи')
            break

